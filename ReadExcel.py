import openpyxl

wb = openpyxl.load_workbook(r"C:\Sample.xlsx")

sheet = wb.active
print(wb.sheetnames)

print('rows count', sheet.max_row)
print('columns count', sheet.max_column)
row = sheet.max_row
col = sheet.max_column

# for i in range(1, row+1):
#     for j in range(1, col+1):
#         cell_obj = sheet.cell(row= i,column= j)
#         print(cell_obj.value)

eng = {}
frn = {}
itl = {}
for i in range(2, row+1):
    key = sheet.cell(row=i, column=1).value
    eng[key] = sheet.cell(row=i, column=2).value
    frn[key] = sheet.cell(row=i, column=3).value
    itl[key] = sheet.cell(row=i, column=4).value

print(eng)
print(frn)
print(itl)
